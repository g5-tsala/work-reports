"""Checks embutidos na geradora.

A planilha se autoconfere: cada uma destas celulas e uma diferenca que deveria
dar zero. Elas sao extraidas como dado, e quem decide se o build para e a etapa
de validacao (`core/validacao.py`), conforme `docs/validacao.md` §2.
"""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from core import config
from core.planilha import numero

#: `(aba, intervalo, o que a diferenca compara)`.
INTERVALOS = (
    ("CEO-Dashboard", "C2:L3", "AUM, run rate e ROA contra a soma dos officers"),
    ("Dashboard", "K16:L32", "captacao cliente contra a soma das quebras"),
    ("aum_receita", "AG6:AG50", "NET e receita contra a soma por tipo de veiculo"),
    ("net_in_out", "R7:S90", "IN/OUT/NET contra a soma dos tipos e o io_portfolios"),
    ("io_grupos", "Q5:Q7", "soma mensal contra o NET e contra o YTD"),
    ("resumo", "AC21:AC24", "total por categoria contra o total por grupo"),
)


def extrair(ctx) -> list[dict[str, Any]]:
    resultados = []
    for aba, intervalo, descricao in INTERVALOS:
        ws = ctx.pl.aba(aba)
        for linha in ws[intervalo]:
            for celula in linha:
                valor = numero(celula.value)
                if valor is None:
                    continue
                referencia = f"{get_column_letter(celula.column)}{celula.row}"
                resultados.append(
                    {
                        "origem": f"{aba}!{referencia}",
                        "descricao": descricao,
                        "valor": valor,
                        "ok": abs(valor) <= config.TOLERANCIA_CHECK,
                    }
                )
    return resultados
