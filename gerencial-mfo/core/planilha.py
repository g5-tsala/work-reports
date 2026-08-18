"""Acesso a planilha: abertura, leitura de intervalos e limpeza de valores.

A planilha mensal e um snapshot com valores colados, entao tudo aqui e leitura
pura (`data_only=True`). Nenhum calculo de negocio vive neste modulo.

Regra de limpeza: `numero()` e `texto()` sao funcoes distintas de proposito.
Traco (`-`) e `TBD` sao "sem valor" quando o campo e numerico, mas sao conteudo
legitimo quando o campo e texto — o officer dos Fdos Alocacao e literalmente
`-`, e converte-lo para nulo apaga uma linha inteira da CEO-Dashboard.
"""

from __future__ import annotations

import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

#: Valores de erro do Excel, na grafia PT-BR e EN-US.
ERROS_EXCEL = frozenset(
    {
        "#DIV/0!",
        "#N/A",
        "#N/D",
        "#NAME?",
        "#NOME?",
        "#NULL!",
        "#NULO!",
        "#NUM!",
        "#REF!",
        "#VALOR!",
        "#VALUE!",
    }
)


def eh_erro(valor: Any) -> bool:
    """`True` se a celula carrega um erro do Excel."""
    return isinstance(valor, str) and valor.strip() in ERROS_EXCEL


def numero(valor: Any) -> float | None:
    """Converte para float. Erro, texto, traco, `TBD` e vazio viram `None`."""
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    return None


def inteiro(valor: Any) -> int | None:
    n = numero(valor)
    return None if n is None else round(n)


def texto(valor: Any) -> str | None:
    """Converte para string. Erro do Excel e vazio viram `None`; traco sobrevive."""
    if valor is None:
        return None
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, str):
        limpo = valor.strip()
        if not limpo or limpo in ERROS_EXCEL:
            return None
        return limpo
    return str(valor)


def mes(valor: Any) -> str | None:
    """Normaliza uma celula de data para `YYYY-MM`."""
    if isinstance(valor, (datetime, date)):
        return f"{valor.year:04d}-{valor.month:02d}"
    if isinstance(valor, str):
        limpo = valor.strip()
        if len(limpo) >= 7 and limpo[4] == "-":
            return limpo[:7]
    return None


#: Simbolos que a conversao para ASCII descartaria, criando colisao de chave:
#: sem isso `AUM (R$)` e `Δ AUM` virariam ambos `aum`.
SIMBOLOS_NO_ROTULO = (("US$", " usd "), ("R$", " rs "), ("%", " pct "), ("Δ", " delta "))


def chave(rotulo: str | None) -> str | None:
    """Slug ASCII estavel a partir de um rotulo da planilha.

    `"Receita Mens. (R$)"` -> `"receita_mens_rs"`. Serve para o template achar
    uma linha pelo nome sem depender da posicao dela na aba.
    """
    if not rotulo:
        return None
    for simbolo, substituto in SIMBOLOS_NO_ROTULO:
        rotulo = rotulo.replace(simbolo, substituto)
    sem_acento = unicodedata.normalize("NFKD", rotulo).encode("ascii", "ignore").decode()
    partes = [p for p in "".join(c if c.isalnum() else " " for c in sem_acento).split()]
    return "_".join(partes).lower() or None


class Planilha:
    """Wrapper de leitura sobre o snapshot mensal."""

    def __init__(self, caminho: Path):
        self.caminho = Path(caminho)
        self.wb = load_workbook(self.caminho, data_only=True)

    def fechar(self) -> None:
        self.wb.close()

    # -- acesso bruto ----------------------------------------------------

    def aba(self, nome: str):
        if nome not in self.wb.sheetnames:
            raise KeyError(f"aba '{nome}' nao existe em {self.caminho.name}")
        return self.wb[nome]

    def celula(self, aba: str, ref: str) -> Any:
        return self.aba(aba)[ref].value

    def bloco(self, aba: str, lin_ini: int, col_ini: int, lin_fim: int, col_fim: int) -> list[list[Any]]:
        ws = self.aba(aba)
        return [
            [ws.cell(linha, coluna).value for coluna in range(col_ini, col_fim + 1)]
            for linha in range(lin_ini, lin_fim + 1)
        ]

    def linha(self, aba: str, linha: int, col_ini: int, col_fim: int) -> list[Any]:
        ws = self.aba(aba)
        return [ws.cell(linha, coluna).value for coluna in range(col_ini, col_fim + 1)]

    def coluna(self, aba: str, coluna: int, lin_ini: int, lin_fim: int) -> list[Any]:
        ws = self.aba(aba)
        return [ws.cell(linha, coluna).value for linha in range(lin_ini, lin_fim + 1)]

    # -- nomes definidos -------------------------------------------------

    def limites_nome(self, nome: str) -> tuple[str, int, int, int, int]:
        """`(aba, lin_ini, col_ini, lin_fim, col_fim)` de um nome definido."""
        definicao = self.wb.defined_names.get(nome)
        if definicao is None:
            raise KeyError(f"nome definido '{nome}' nao existe em {self.caminho.name}")
        destinos = list(definicao.destinations)
        if not destinos:
            raise ValueError(f"nome definido '{nome}' nao aponta para um intervalo")
        aba, ref = destinos[0]
        col_ini, lin_ini, col_fim, lin_fim = range_boundaries(ref)
        return aba, lin_ini, col_ini, lin_fim, col_fim

    def bloco_nome(self, nome: str) -> list[list[Any]]:
        aba, lin_ini, col_ini, lin_fim, col_fim = self.limites_nome(nome)
        return self.bloco(aba, lin_ini, col_ini, lin_fim, col_fim)

    def valor_nome(self, nome: str) -> Any:
        aba, lin_ini, col_ini, _, _ = self.limites_nome(nome)
        return self.aba(aba).cell(lin_ini, col_ini).value


def limites_ref(ref: str) -> tuple[int, int, int, int]:
    """`(lin_ini, col_ini, lin_fim, col_fim)` de uma referencia tipo `$C$31:$O$63`."""
    col_ini, lin_ini, col_fim, lin_fim = range_boundaries(ref)
    return lin_ini, col_ini, lin_fim, col_fim
